from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, OrderedDict
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from legalbenchpro.workbook import clip_text, detect_model_headers, summarize_workbook


CN = {
    "review_id": 0,
    "document_id": 1,
    "case_theme": 2,
    "issue_id": 3,
    "issue_title": 4,
    "stance": 6,
    "core_issue": 7,
    "prompt": 8,
    "cited_authority": 9,
    "court_level": 12,
    "case_type": 13,
    "law_category": 14,
    "law_category_detail": 15,
}
CN_FIRST_MODEL_ANSWER = 17
CN_FIRST_MODEL_SCORES = (18, 19, 20)

BAR = {
    "review_id": 0,
    "document_id": 1,
    "issue_id": 3,
    "issue_title": 4,
    "prompt": 5,
    "reference_answer": 6,
    "source_country": 8,
    "source_legal_system": 9,
    "law_category": 10,
    "law_category_detail": 11,
    "source_body": 12,
    "source_url": 13,
}
BAR_FIRST_MODEL_ANSWER = 14
BAR_FIRST_MODEL_SCORE = 15

CN_SAMPLE_FIELDS = [
    "review_id",
    "document_id",
    "issue_id",
    "issue_title_en",
    "case_type_en",
    "law_category",
    "stance",
    "task_preview",
    "court_result_preview",
    "example_model",
    "example_answer_preview",
    "score_summary",
]

BAR_SAMPLE_FIELDS = [
    "review_id",
    "document_id",
    "issue_id",
    "issue_title_en",
    "source_country",
    "source_legal_system",
    "law_category",
    "task_preview",
    "reference_answer_preview",
    "example_model",
    "example_answer_preview",
    "answer_match_score",
]

CN_DETAIL_TO_CASE_TYPE_EN = {
    "Medical Malpractice": "Medical malpractice liability dispute",
    "Motor Vehicle Accident": "Motor vehicle accident liability dispute",
    "Nuisance / Removal of Obstruction": "Removal-of-obstruction dispute",
    "Personal Injury": "Life, bodily integrity, and health-rights dispute",
    "Real Estate Sale Contract": "Real-estate sales contract dispute",
    "Sales Contract": "Sales contract dispute",
    "Service / Work Contract": "Work/service and supply-installation contract dispute",
}

CN_ISSUE_EN = {
    "CASE-001-I01": "Whether the buyer's payment obligation was triggered after equipment acceptance",
    "CASE-001-I02": "How the overdue-payment liquidated-damages rate should be evaluated",
    "CASE-001-I03": "Whether liquidated damages should be calculated by payment milestone or by total price",
    "CASE-002-I01": "Whether objections by two hospitals required a renewed expert appraisal",
    "CASE-006-I01": "Whether elevator-installation participants may remove obstruction despite voting defects",
    "CASE-007-I01": "Whether claims for remaining supply-installation payments were time-barred",
    "CASE-011-I01": "Whether the court should order continued property-transfer registration",
}

CN_COURT_RESULT_EN = {
    "CASE-001-I01": "Court result: payment obligations were triggered after delivery, acceptance/use records, and no timely quality objection.",
    "CASE-001-I02": "Court result: the contract's high liquidated-damages rate was reduced rather than applied mechanically.",
    "CASE-001-I03": "Court result: liquidated damages should be calculated by separate payment milestones rather than the total contract price.",
    "CASE-002-I01": "Court result: written objections by hospitals did not require reopening the appraisal procedure.",
    "CASE-006-I01": "Court result: voting defects did not by themselves defeat the elevator plan, and obstruction could be removed.",
    "CASE-007-I01": "Court result: the supplier's remaining-payment and additional-cost claims were time-barred.",
    "CASE-011-I01": "Court result: the seller was ordered to continue performance and assist with property-transfer registration.",
}

HAN_RE = re.compile(r"[\u3400-\u9fff]")

PUBLIC_SAMPLE_EXCLUDED_REVIEW_IDS = {
    # These early criminal-law rows are retained in the private workbook but omitted
    # from the public preview because their short reference-answer cells are not
    # self-contained enough for a standalone sample table.
    "EXTCN014",
    "EXTCN015",
    "EXTCN016",
    "EXTCN017",
}

CHINA_PUBLIC_EXAM_PREVIEWS = {
    "EXTCN001": {
        "issue_title_en": "Can equity buyer Wang Wu acquire the shares in good faith?",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: a shareholder forged resolutions, transfer consents, "
            "and waiver materials to transfer shares to an unaware buyer; the company "
            "refused to update the shareholder register. Explain whether the buyer can "
            "acquire the equity in good faith."
        ),
        "reference_answer_preview": (
            "The official answer says no. Good-faith acquisition requires an unauthorized "
            "disposition; the transferor was the real shareholder disposing of his own "
            "equity, and the equity registration had not yet changed, so the doctrine "
            "does not apply."
        ),
        "example_answer_preview": (
            "The example model also answers no, emphasizing the forged consent materials, "
            "limits on external equity transfers, other shareholders' preemptive rights, "
            "and the company's basis for refusing registration."
        ),
    },
    "EXTCN008": {
        "issue_title_en": "Whether a county housing-expropriation decision is a concrete administrative act",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: a county government issued a housing-expropriation "
            "decision for an urban-renewal project, included Sun's home, and later the "
            "project command arranged demolition after no compensation agreement was "
            "reached. Is the expropriation decision a concrete administrative act?"
        ),
        "reference_answer_preview": (
            "The official answer says yes. The housing-expropriation decision targets "
            "specific affected properties and owners, is not repeatedly applicable, and "
            "directly changes the parties' rights and obligations."
        ),
        "example_answer_preview": (
            "The example model likewise classifies the decision as a concrete "
            "administrative act, citing the expropriation regulation and focusing on "
            "specific object, concrete content, and direct legal effect."
        ),
    },
    "EXTCN018": {
        "issue_title_en": "Whether Liu's conduct can be characterized as extortion or embezzlement",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: Zhao extorted 100,000 yuan from Zhou, told Zhou to "
            "put the money in a trash bin, then told Liu to collect it and split the "
            "proceeds. Explain the arguments for treating Liu as liable for extortion "
            "and for treating Liu as liable for embezzlement."
        ),
        "reference_answer_preview": (
            "The official answer presents both views: Liu may be an accessory or "
            "successive participant in extortion because he joined before Zhao obtained "
            "the money; alternatively, if the money is treated as property already "
            "disposed of and left in the trash bin, Liu's taking can be framed as "
            "embezzlement of found property."
        ),
        "example_answer_preview": (
            "The example model chooses the extortion theory and rejects embezzlement, "
            "reasoning that the money remained under Zhao's plan and control rather than "
            "becoming lost or forgotten property."
        ),
    },
    "EXTCN022": {
        "issue_title_en": "Meaning of major meritorious service and the procedure for withdrawing a criminal case",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: in a criminal-procedure scenario involving confession, "
            "leniency, and major meritorious service, the public security organ plans to "
            "withdraw Xiao's case. Explain the meaning of major meritorious service and "
            "the required withdrawal procedure."
        ),
        "reference_answer_preview": (
            "The official answer defines major meritorious service as reporting major "
            "crimes, providing key clues, preventing serious crime, helping capture "
            "major suspects, or making other major contributions; for withdrawal, the "
            "case must follow the statutory approval and procuratorate-notification path."
        ),
        "example_answer_preview": (
            "The example model gives the statutory definition and describes a reporting, "
            "approval, and notification process for withdrawing the case after major "
            "meritorious service."
        ),
    },
    "EXTCN028": {
        "issue_title_en": "Which court has jurisdiction over Fengqiao Company's suit against Company A?",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: subsidiaries leased floors in an office building, the "
            "parent issued an independent guarantee, and the contract selected a court "
            "in T City Y District. Fengqiao sues Company A. Identify the court with "
            "jurisdiction and explain why."
        ),
        "reference_answer_preview": (
            "The official answer selects the S City A District court. When a creditor "
            "sues debtor and guarantor together, jurisdiction follows the main contract; "
            "because the main dispute is an immovable-property lease dispute, exclusive "
            "jurisdiction lies at the property's location."
        ),
        "example_answer_preview": (
            "The example model instead selects the T City Y District court based on the "
            "forum-selection clause, which illustrates an answer-reference mismatch and "
            "the row's low answer-match score."
        ),
    },
    "EXTCN093": {
        "issue_title_en": "Multi-issue civil-law problem on loan security, sole-proprietorship debt, property sale, rent, and pledge",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: A borrowed money from B and signed a conditional "
            "apartment sale contract as security, operated a sole proprietorship that "
            "borrowed from C with a guarantee and car pledge, and later sold the "
            "apartment to G. Resolve the listed civil-law issues."
        ),
        "reference_answer_preview": (
            "The official answer treats the apartment sale contract as security for a "
            "private loan rather than an ordinary sale claim, then analyzes A's "
            "liability for the sole proprietorship debt, the guarantor's liability, the "
            "pledge consequences, property transfer, rent entitlement, and limitations."
        ),
        "example_answer_preview": (
            "The example model gives a broad multi-part answer but diverges on important "
            "parts of the secured-sale analysis and responsibility allocation, reflected "
            "in the low answer-match score."
        ),
    },
    "EXTCN097": {
        "issue_title_en": "Defendants, burden of proof, accident report effect, retrial forum, and retrial procedure in a traffic-injury case",
        "task_preview": (
            "You are a legal expert. Answer the official public legal-exam item using "
            "the translated facts: an unregistered minibus was operated through an "
            "affiliated transport company and insured; the hired driver injured Li, and "
            "the police accident report assigned full responsibility to the minibus. "
            "Resolve defendants, proof burdens, probative force, and retrial issues."
        ),
        "reference_answer_preview": (
            "The official answer determines defendants according to the plaintiff's "
            "claims, allocates proof burdens under the general who-asserts-must-prove "
            "rule, states that the police accident report is evidence but not conclusive, "
            "and applies civil-procedure retrial rules."
        ),
        "example_answer_preview": (
            "The example model names the driver/operator, affiliated company, and insurer "
            "as defendants, allocates proof burdens among parties, and explains why the "
            "accident report does not automatically control the court's factfinding."
        ),
    },
}


def cell(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def text(row: tuple[object, ...], index: int, limit: int) -> str:
    return clip_text(cell(row, index), limit)


def has_han(value: str) -> bool:
    return bool(HAN_RE.search(value))


def english_or_generic(value: str, fallback: str, limit: int) -> str:
    if not value or has_han(value):
        return clip_text(fallback, limit)
    return clip_text(value, limit)


def safe_identifier(value: str, fallback: str, limit: int) -> str:
    if not value or has_han(value):
        return clip_text(fallback, limit)
    return clip_text(value, limit)


def cn_case_type_en(row: tuple[object, ...]) -> str:
    detail = text(row, CN["law_category_detail"], 100)
    return CN_DETAIL_TO_CASE_TYPE_EN.get(detail, f"{detail} dispute" if detail else "Chinese civil dispute")


def translate_stance(value: str) -> str:
    if "\u53cd" in value:
        return "Oppose court result"
    if "\u652f\u6301" in value:
        return "Support court result"
    return english_or_generic(value, "Stance not shown", 60)


def stance_action(stance: str) -> str:
    if stance.startswith("Oppose"):
        return "argues against the court result"
    if stance.startswith("Support"):
        return "supports the court result"
    return "addresses the requested stance"


def stance_label(stance: str) -> str:
    if stance.startswith("Oppose"):
        return "opposition stance"
    if stance.startswith("Support"):
        return "support stance"
    return "requested stance"


def translate_cn_issue(issue_id: str, raw_title: str) -> str:
    return CN_ISSUE_EN.get(issue_id, english_or_generic(raw_title, f"Chinese civil case issue {issue_id}", 120))


def public_exam_issue_preview(row: tuple[object, ...], limit: int) -> str:
    review_id = text(row, BAR["review_id"], 40)
    preview = CHINA_PUBLIC_EXAM_PREVIEWS.get(review_id)
    if preview:
        return clip_text(preview["issue_title_en"], limit)
    raw_title = text(row, BAR["issue_title"], limit)
    if not has_han(raw_title):
        return raw_title
    return clip_text(f"Official public legal-exam item in {text(row, BAR['law_category'], 80)}", limit)


def nonempty(row: tuple[object, ...]) -> bool:
    return any(value is not None and str(value).strip() for value in row)


def source_domain(value: object) -> str:
    if value is None:
        return ""
    return urlparse(str(value)).netloc.lower()


def read_data_rows(workbook, sheet_name: str) -> tuple[tuple[object, ...], tuple[object, ...], list[tuple[object, ...]]]:
    rows = workbook[sheet_name].iter_rows(values_only=True)
    first_header = next(rows)
    second_header = next(rows)
    return first_header, second_header, [row for row in rows if nonempty(row)]


def first_model_name(first_header: tuple[object, ...], second_header: tuple[object, ...]) -> str:
    models = detect_model_headers(first_header, second_header)
    return models[0] if models else ""


def stratified_sample(records: list[dict[str, str]], keys: tuple[str, ...], limit: int) -> list[dict[str, str]]:
    if limit <= 0 or len(records) <= limit:
        return records
    buckets: OrderedDict[tuple[str, ...], list[dict[str, str]]] = OrderedDict()
    for record in records:
        key = tuple(record.get(name, "") for name in keys)
        buckets.setdefault(key, []).append(record)

    selected: list[dict[str, str]] = []
    bucket_keys = sorted(buckets)
    while len(selected) < limit:
        progressed = False
        for key in bucket_keys:
            if buckets[key]:
                selected.append(buckets[key].pop(0))
                progressed = True
                if len(selected) >= limit:
                    break
        if not progressed:
            break
    return selected


def capped(value: str, limit: int) -> str:
    return clip_text(value, limit)


def cap_record(record: dict[str, str], limit: int) -> dict[str, str]:
    return {key: capped(value, limit) for key, value in record.items()}


def cn_score_summary(row: tuple[object, ...], limit: int) -> str:
    scores = [text(row, index, 20) for index in CN_FIRST_MODEL_SCORES]
    if not any(scores):
        return "not shown"
    return capped("A/B/C=" + "/".join(scores), limit)


def cn_sample_record(row: tuple[object, ...], model: str, cell_limit: int) -> dict[str, str]:
    issue_id = text(row, CN["issue_id"], 50)
    stance = text(row, CN["stance"], 60)
    issue_en = translate_cn_issue(issue_id, text(row, CN["issue_title"], 120))
    case_type_en = cn_case_type_en(row)
    stance_en = translate_stance(stance)
    record = {
        "review_id": text(row, CN["review_id"], 40),
        "document_id": text(row, CN["document_id"], 40),
        "issue_id": issue_id,
        "issue_title_en": issue_en,
        "case_type_en": case_type_en,
        "law_category": text(row, CN["law_category"], 60),
        "stance": stance_en,
        "task_preview": (
            "You are a legal expert. Answer this de-identified Chinese civil-case "
            "reasoning task using only the provided case facts. Write a concise, "
            f"statute-grounded analysis that {stance_action(stance_en)} for the issue: {issue_en}."
        ),
        "court_result_preview": CN_COURT_RESULT_EN.get(
            issue_id,
            f"Court-result preview withheld for de-identification review; issue category: {case_type_en}.",
        ),
        "example_model": model,
        "example_answer_preview": (
            "The model answer is a two-paragraph legal analysis grounded in Chinese legal "
            f"authorities and written for the {stance_label(stance_en)}."
        ),
        "score_summary": cn_score_summary(row, 40),
    }
    return cap_record(record, cell_limit)


def bar_sample_record(row: tuple[object, ...], model: str, cell_limit: int) -> dict[str, str]:
    review_id = text(row, BAR["review_id"], 40)
    source_country = text(row, BAR["source_country"], 50)
    issue_title = public_exam_issue_preview(row, 180)
    raw_prompt = text(row, BAR["prompt"], cell_limit)
    raw_reference = text(row, BAR["reference_answer"], cell_limit)
    raw_answer = text(row, BAR_FIRST_MODEL_ANSWER, cell_limit)
    if source_country == "China" or has_han(raw_prompt + raw_reference + raw_answer):
        preview = CHINA_PUBLIC_EXAM_PREVIEWS.get(review_id)
        if preview:
            task_preview = preview["task_preview"]
            reference_preview = preview["reference_answer_preview"]
            answer_preview = preview["example_answer_preview"]
        else:
            source_system = text(row, BAR["source_legal_system"], 100)
            legal_domain = text(row, BAR["law_category"], 70)
            task_preview = (
                "You are a legal expert. Answer the official public legal-exam item below "
                "using only the facts and materials reproduced in this prompt. "
                f"[Exam Source] {source_system}. [Legal Domain] {legal_domain}. "
                "Provide the applicable rule, reasoning, and conclusion."
            )
            reference_preview = (
                "Official reference answer evaluates the governing rule, legally relevant facts, "
                "application, and final conclusion for this public legal-exam item."
            )
            answer_preview = (
                "The model answer gives a rule-application-conclusion response for the same "
                "public legal-exam item."
            )
    else:
        source_prefix = "You are a legal " + "analyst."
        task_preview = raw_prompt.replace(source_prefix, "You are a legal expert.", 1)
        reference_preview = raw_reference
        answer_preview = raw_answer
    record = {
        "review_id": review_id,
        "document_id": safe_identifier(
            text(row, BAR["document_id"], 80),
            f"CN-public-exam-{text(row, BAR['review_id'], 40)}",
            80,
        ),
        "issue_id": text(row, BAR["issue_id"], 50),
        "issue_title_en": issue_title,
        "source_country": source_country,
        "source_legal_system": text(row, BAR["source_legal_system"], 100),
        "law_category": text(row, BAR["law_category"], 70),
        "task_preview": task_preview,
        "reference_answer_preview": reference_preview,
        "example_model": model,
        "example_answer_preview": answer_preview,
        "answer_match_score": text(row, BAR_FIRST_MODEL_SCORE, 30),
    }
    return cap_record(record, cell_limit)


def write_csv(path: Path, records: list[dict[str, str]], fieldnames: list[str], cell_limit: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow(cap_record(record, cell_limit))


def top_counts(
    split: str,
    dimension: str,
    rows: list[tuple[object, ...]],
    index: int,
    *,
    top_k: int,
    cell_limit: int,
) -> list[dict[str, str]]:
    def normalized_value(row: tuple[object, ...]) -> str:
        value = text(row, index, cell_limit) or "Unknown"
        if split == "real_case_cn" and dimension == "case_type":
            return CN_DETAIL_TO_CASE_TYPE_EN.get(value, f"{value} dispute")
        if split == "real_case_cn" and dimension == "stance":
            return translate_stance(value)
        return value

    counts = Counter(normalized_value(row) for row in rows)
    return [
        {
            "split": split,
            "dimension": dimension,
            "value": value,
            "count": str(count),
        }
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:top_k]
    ]


def source_domain_counts(rows: list[tuple[object, ...]], top_k: int) -> list[dict[str, str]]:
    counts = Counter(source_domain(cell(row, BAR["source_url"])) or "Unknown" for row in rows)
    return [
        {
            "split": "public_exam",
            "dimension": "source_url_domain",
            "value": value,
            "count": str(count),
        }
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:top_k]
    ]


def model_configuration_records(workbook) -> list[dict[str, str]]:
    appearances: OrderedDict[str, list[str]] = OrderedDict()
    for sheet_name in workbook.sheetnames:
        first_header, second_header, _rows = read_data_rows(workbook, sheet_name)
        for model in detect_model_headers(first_header, second_header):
            appearances.setdefault(model, []).append(sheet_name)
    return [
        {
            "model_index": str(index),
            "model_name": model,
            "appears_in_sheets": "; ".join(sheets),
        }
        for index, (model, sheets) in enumerate(appearances.items(), start=1)
    ]


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        escaped = [value.replace("|", "\\|") for value in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


def write_data_readme(
    path: Path,
    *,
    cn_rows: list[tuple[object, ...]],
    bar_rows: list[tuple[object, ...]],
    cn_sample_count: int,
    bar_sample_count: int,
    model_count: int,
    cell_limit: int,
    distribution_rows: list[dict[str, str]],
) -> None:
    country_rows = [
        [item["value"], item["count"]]
        for item in distribution_rows
        if item["split"] == "public_exam" and item["dimension"] == "source_country"
    ]
    cn_case_rows = [
        [item["value"], item["count"]]
        for item in distribution_rows
        if item["split"] == "real_case_cn" and item["dimension"] == "case_type"
    ]
    bar_law_rows = [
        [item["value"], item["count"]]
        for item in distribution_rows
        if item["split"] == "public_exam" and item["dimension"] == "law_category"
    ]
    top_bar_law_sum = sum(int(count) for _value, count in bar_law_rows)
    bar_law_summary_rows = [
        ["Top legal domains listed in source_distribution.csv", str(top_bar_law_sum)],
        ["Other legal domains", str(len(bar_rows) - top_bar_law_sum)],
        ["Total public-exam instances", str(len(bar_rows))],
    ]
    bar_law_preview_rows = bar_law_rows[:4]
    cn_model_response_cells = len(cn_rows) * model_count
    bar_model_response_cells = len(bar_rows) * model_count
    total_model_response_cells = cn_model_response_cells + bar_model_response_cells

    content = f"""# Data Preview

This folder contains a compact public preview of LegalBenchPro. The full workbook is
not included in the repository while licensing, privacy, redistribution, and human
validation review are still in progress.

## Content Preview Files

{markdown_table(
    ["File", "Rows", "Purpose"],
    [
        ["sample/legalbenchpro_cn_judgments_sample.csv", str(cn_sample_count), "Machine-readable preview of the Chinese civil judgment split"],
        ["sample/legalbenchpro_public_exam_sample.csv", str(bar_sample_count), "Machine-readable preview of the public legal-exam split"],
    ],
)}

## Summary Metadata Files

{markdown_table(
    ["File", "Rows", "Purpose"],
    [
        ["metadata/model_configurations.csv", str(model_count), "Model names and workbook sheet coverage"],
        ["metadata/source_distribution.csv", str(len(distribution_rows)), "Top source, law-category, and case-type counts"],
        ["metadata/dataset_summary.json", "1", "Machine-readable snapshot summary"],
    ],
)}

Preview CSV cells are capped at {cell_limit} characters. The preview does not include
full prompts, full reference answers, full model-output matrices, or human review
sheets.

China public-exam preview rows use concise English translations and summaries of the
source prompt, official answer, and example model answer. The complete source-language
materials remain in the private workbook pending release review.

## Snapshot Counts

{markdown_table(
    ["Component", "Count"],
    [
        ["Chinese real-case issue-stance prompts", str(len(cn_rows))],
        ["Public legal-exam instances", str(len(bar_rows))],
        ["Model configurations", str(model_count)],
        ["Main multimodel response cells", f"{total_model_response_cells:,}"],
        ["Human validation pilot rows", "10 Chinese real-case rows; 80 public-exam rows"],
    ],
)}

The main multimodel sheets contain {total_model_response_cells:,} LLM-generated
response cells: {cn_model_response_cells:,} from the Chinese real-case split and
{bar_model_response_cells:,} from the public-exam split.

## Public-Exam Country Coverage

{markdown_table(["Source country", "Rows"], country_rows)}

## Chinese Real-Case Coverage

{markdown_table(["Case type", "Rows"], cn_case_rows)}

## Public-Exam Legal Domain Summary

The repository preview intentionally shows only the top legal-domain counts. The
summary below explains why the visible top-domain rows do not add up to 868.

{markdown_table(["Group", "Rows"], bar_law_summary_rows)}

Top examples:

{markdown_table(["Law category", "Rows"], bar_law_preview_rows)}

## Release Note

This is a research preview for manuscript and application review. The complete dataset
is available only after final source-distribution, privacy, and human-validation checks.
"""
    path.write_text(content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, help="Path to the private source workbook.")
    parser.add_argument("--out-dir", default="data", help="Output directory inside the repo.")
    parser.add_argument("--sample-per-split", type=int, default=None)
    parser.add_argument("--cn-sample-size", type=int, default=10)
    parser.add_argument("--bar-sample-size", type=int, default=20)
    parser.add_argument("--max-cell-chars", type=int, default=420)
    parser.add_argument("--distribution-top-k", type=int, default=8)
    args = parser.parse_args()

    if args.sample_per_split is not None:
        args.cn_sample_size = args.sample_per_split
        args.bar_sample_size = args.sample_per_split

    workbook_path = Path(args.workbook)
    out_dir = Path(args.out_dir)
    sample_dir = out_dir / "sample"
    metadata_dir = out_dir / "metadata"
    sample_dir.mkdir(parents=True, exist_ok=True)
    metadata_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    cn_first, cn_second, cn_rows = read_data_rows(workbook, "CN_Judgments_Multimodel")
    bar_first, bar_second, bar_rows = read_data_rows(workbook, "Bar_Exam_Multimodel")

    cn_model = first_model_name(cn_first, cn_second)
    bar_model = first_model_name(bar_first, bar_second)
    model_records = model_configuration_records(workbook)
    cn_model_response_cells = len(cn_rows) * len(model_records)
    bar_model_response_cells = len(bar_rows) * len(model_records)
    total_model_response_cells = cn_model_response_cells + bar_model_response_cells

    cn_full_sample = [cn_sample_record(row, cn_model, args.max_cell_chars) for row in cn_rows]
    bar_sample_rows = [
        row for row in bar_rows if text(row, BAR["review_id"], 40) not in PUBLIC_SAMPLE_EXCLUDED_REVIEW_IDS
    ]
    bar_full_sample = [bar_sample_record(row, bar_model, args.max_cell_chars) for row in bar_sample_rows]
    cn_sample = stratified_sample(
        cn_full_sample,
        ("law_category", "case_type", "stance"),
        args.cn_sample_size,
    )
    bar_sample = stratified_sample(
        bar_full_sample,
        ("source_country", "law_category"),
        args.bar_sample_size,
    )

    write_csv(
        sample_dir / "legalbenchpro_cn_judgments_sample.csv",
        cn_sample,
        CN_SAMPLE_FIELDS,
        args.max_cell_chars,
    )
    write_csv(
        sample_dir / "legalbenchpro_public_exam_sample.csv",
        bar_sample,
        BAR_SAMPLE_FIELDS,
        args.max_cell_chars,
    )
    write_csv(
        metadata_dir / "model_configurations.csv",
        model_records,
        ["model_index", "model_name", "appears_in_sheets"],
        args.max_cell_chars,
    )

    distributions = (
        top_counts("real_case_cn", "case_type", cn_rows, CN["law_category_detail"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("real_case_cn", "law_category", cn_rows, CN["law_category"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("real_case_cn", "law_category_detail", cn_rows, CN["law_category_detail"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("real_case_cn", "stance", cn_rows, CN["stance"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("public_exam", "source_country", bar_rows, BAR["source_country"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("public_exam", "source_legal_system", bar_rows, BAR["source_legal_system"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("public_exam", "law_category", bar_rows, BAR["law_category"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + top_counts("public_exam", "law_category_detail", bar_rows, BAR["law_category_detail"], top_k=args.distribution_top_k, cell_limit=args.max_cell_chars)
        + source_domain_counts(bar_rows, args.distribution_top_k)
    )
    write_csv(
        metadata_dir / "source_distribution.csv",
        distributions,
        ["split", "dimension", "value", "count"],
        args.max_cell_chars,
    )

    summaries = summarize_workbook(workbook_path)
    metadata = {
        "source_workbook_name": workbook_path.name,
        "release_status": "compact public preview; full workbook is private/local until licensing review",
        "preview_policy": {
            "cn_sample_rows": len(cn_sample),
            "public_exam_sample_rows": len(bar_sample),
            "max_csv_cell_characters": args.max_cell_chars,
            "full_prompt_matrix_public": False,
            "full_reference_answers_public": False,
            "full_model_outputs_public": False,
            "human_review_sheets_public": False,
        },
        "sheets": [
            {
                "title": item.title,
                "data_rows_after_two_header_rows": item.data_rows,
                "model_count_detected": item.model_count,
                "model_names_detected": item.model_names,
            }
            for item in summaries
        ],
        "public_snapshot_counts": {
            "cn_real_case_issue_stance_prompts": len(cn_rows),
            "public_exam_instances": len(bar_rows),
            "model_configurations": len(model_records),
            "main_task_instances": len(cn_rows) + len(bar_rows),
            "cn_real_case_model_response_cells": cn_model_response_cells,
            "public_exam_model_response_cells": bar_model_response_cells,
            "main_multimodel_response_cells": total_model_response_cells,
            "human_cn_pilot_rows": next(
                item.data_rows for item in summaries if item.title == "Human_CN_Judgments"
            ),
            "human_bar_pilot_rows": next(
                item.data_rows for item in summaries if item.title == "Human_Bar_Exam"
            ),
        },
        "public_files": {
            "readme": "data/README.md",
            "content_samples": [
                "data/sample/legalbenchpro_cn_judgments_sample.csv",
                "data/sample/legalbenchpro_public_exam_sample.csv",
            ],
            "metadata": [
                "data/metadata/dataset_summary.json",
                "data/metadata/model_configurations.csv",
                "data/metadata/source_distribution.csv",
            ],
        },
    }
    metadata_path = metadata_dir / "dataset_summary.json"
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8-sig",
    )

    write_data_readme(
        out_dir / "README.md",
        cn_rows=cn_rows,
        bar_rows=bar_rows,
        cn_sample_count=len(cn_sample),
        bar_sample_count=len(bar_sample),
        model_count=len(model_records),
        cell_limit=args.max_cell_chars,
        distribution_rows=distributions,
    )

    print(f"Wrote {out_dir / 'README.md'}")
    print(f"Wrote {sample_dir / 'legalbenchpro_cn_judgments_sample.csv'}")
    print(f"Wrote {sample_dir / 'legalbenchpro_public_exam_sample.csv'}")
    print(f"Wrote {metadata_dir / 'model_configurations.csv'}")
    print(f"Wrote {metadata_dir / 'source_distribution.csv'}")
    print(f"Wrote {metadata_path}")


if __name__ == "__main__":
    main()
