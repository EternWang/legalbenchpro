from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def clip_text(value: Any, limit: int = 320) -> str:
    """Return a compact, public-display-safe excerpt."""
    if value is None:
        return ""
    text = " ".join(str(value).replace("\r", "\n").split())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def detect_model_headers(
    first_header_row: Iterable[Any],
    second_header_row: Iterable[Any] | None = None,
) -> list[str]:
    answer_markers = ("AI Answer", "AI回答", "Model Answer")
    models: list[str] = []
    second_values = list(second_header_row or [])
    for index, value in enumerate(first_header_row):
        if isinstance(value, str):
            name = value.strip()
            if not name:
                continue
            if second_values:
                paired_header = str(second_values[index] or "") if index < len(second_values) else ""
                if not any(marker in paired_header for marker in answer_markers):
                    continue
            elif "Sample" in name or "Info" in name or "基础信息" in name:
                continue
            if name:
                models.append(name)
    return models


@dataclass(frozen=True)
class SheetSummary:
    title: str
    data_rows: int
    model_count: int
    model_names: list[str]


def summarize_workbook(path: str | Path) -> list[SheetSummary]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    summaries: list[SheetSummary] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        first_header = next(rows, ()) or ()
        second_header = next(rows, ()) or ()
        data_rows = sum(1 for row in rows if any(value is not None for value in row))
        models = detect_model_headers(first_header, second_header)
        summaries.append(
            SheetSummary(
                title=sheet.title,
                data_rows=data_rows,
                model_count=len(models),
                model_names=models,
            )
        )
    return summaries
